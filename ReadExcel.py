#read data from excel file testdata.xlsx
import openpyxl
path="E:\WORKSPACE\SeleniumTrainingFrom23Sept\Resource\TestData.xlsx"
workbook= openpyxl.load_workbook(path)
sheet= workbook.active #workbook.get_sheet_by_name("TestData")
rows= sheet.max_row
cols= sheet.max_column

print("Rows: ",rows)
print("Column : ",cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="    ")

    print()